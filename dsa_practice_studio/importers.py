import csv
from pathlib import Path

from dsa_practice_studio.utils import slugify


def _normalize_row(row):
    return [str(cell).strip() if cell is not None else "" for cell in row]


def _looks_like_header(row):
    tokens = [cell.lower() for cell in row if cell]
    has_unit = any("unit" in token for token in tokens)
    has_chapter = any("chapter" in token for token in tokens)
    has_title = any(token in {"question", "title", "problem"} or "question" in token for token in tokens)
    return has_unit and has_chapter and has_title


def parse_tabular_rows(rows, default_unit=None, default_chapter=None, has_header=None):
    lessons = []
    order = 1
    if rows and has_header is None:
        has_header = _looks_like_header(_normalize_row(rows[0]))
    start_index = 1 if has_header else 0

    for row in rows[start_index:]:
        row = _normalize_row(row)
        if not row or all(not cell for cell in row):
            continue
        while len(row) < 7:
            row.append("")

        unit = row[0] or default_unit or "Imported"
        chapter = row[1] or default_chapter or "General"
        title = row[2].strip()
        if not title:
            continue
        leetcode_url = row[3].strip()
        youtube_url = row[4].strip()
        note = row[5].strip()
        difficulty = row[6].strip()

        lesson_id = f"sheet-{slugify(unit)}-{slugify(chapter)}-{slugify(title)}"
        lessons.append(
            {
                "id": lesson_id,
                "title": title,
                "unit": unit,
                "chapter": chapter,
                "kind": "sheet",
                "source": "import",
                "leetcode_url": leetcode_url,
                "youtube_url": youtube_url,
                "resource_url": "",
                "difficulty": difficulty,
                "notes": note,
                "order": order,
            }
        )
        order += 1

    return lessons


def parse_csv_lessons(path, default_unit=None, default_chapter=None, has_header=None, delimiter=","):
    with open(path, "r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        rows = list(reader)
    return parse_tabular_rows(rows, default_unit, default_chapter, has_header)


def parse_csv_text(text, default_unit=None, default_chapter=None, has_header=None, delimiter=","):
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    rows = list(reader)
    return parse_tabular_rows(rows, default_unit, default_chapter, has_header)


def parse_xlsx_lessons(path, default_unit=None, default_chapter=None, has_header=None, sheet_name=None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Missing openpyxl. Install it to import Excel files.") from exc

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    return parse_tabular_rows(rows, default_unit, default_chapter, has_header)
